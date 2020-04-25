#################################################################################################
import openpyxl as ExcelHandler
#################################################################################################


#################################################################################################
#################################################################################################
##### GLOBAL FUNCTIONS #####
#################################################################################################
#################################################################################################

#################################################################################################
### Init Handler ###
#################################################################################################
## A function to display the welcome screen of the system ##
def WelcomeScreen():
	ReadFromExcel()
	print("")
	print("				|*********************************************************************|")
	print("				|*********************************************************************|")
	print("				|*                                                                   *|")
	print("				|*               Welcome to Hospital Management System               *|")
	print("				|*                                                                   *|")
	print("				|* This system is intended to:                                       *|")
	print("				|* A) As an admin:                                                   *|")
	print("				|*    1) Add, delete, edit and display the information of patients   *|")
	print("				|*    2) Add, delete, edit and display the information of doctors    *|")
	print("				|*    3) Book, edit and cancel appointments                          *|")
	print("				|* B) As a user:                                                     *|")
	print("				|*    1) View all hospital departments                               *|")
	print("				|*    2) View all hospital doctors in details                        *|")
	print("				|*    3) View all hospital patients in details                       *|")
	print("				|*    4) View detailed information about one selected patient        *|")
	print("				|*    5) View detailed information about one selected appointment    *|")
	print("				|*                                                                   *|")
	print("				|*********************************************************************|")
	print("				|*********************************************************************|")
#################################################################################################

#################################################################################################
### Mode Selector Handlers ###
#################################################################################################
## A function to ask for the selected mode ##
def AskForMode():
	print("")
	# Asking for the selected mode (Admin Mode or User Mode) #
	print("Please choose the system mode:")
	print("0- Exit.")
	print("1- Admin Mode.")
	print("2- User Mode.")
	Mode = str(input(">> "))
	print("")
	return Mode
#################################################################################################
## A function to ask for the password for the admin mode ##	
def AskForPassword():
	# Asking the user to enter the admin's mode password 
	Password = str(input("Please enter the admin's mode 4-digit password: "))
	print("")
	return Password
#################################################################################################
## A function to ask for the selected mode in the admin mode ##		
def AskForAdminMode():
	# Asking for the selected mode in Admin Mode (Manage Patients, Manage Doctors or Book an Appointment) #
	print("Please choose the admin's mode required feature:")	
	print("0- Return Back.")
	print("1- Manage Patients.")
	print("2- Manage Doctors.")	
	print("3- Manage Appointments.")	
	AdminMode = str(input(">> "))
	print("")
	return AdminMode
#################################################################################################
## A function to ask for the selected option in the admin mode - manage patients ##
def AskForAdminModeManagePatients():
	# Asking for the selected operation in Admin Mode - Manage Patients (Add, Delete, Edit or Display a Patient) #
	print("Please choose one of the following options:")	
	print("0- Return Back.")
	print("1- Add a Patient.")
	print("2- Delete a Patient.")	
	print("3- Edit a Patient.")	
	print("4- Display a Patient.")	
	print("5- Display All Patients.")	
	PatientOption = str(input(">> "))
	print("")
	return PatientOption
#################################################################################################
## A function to ask for the selected option in the admin mode - manage doctors ##
def AskForAdminModeManageDoctors():
	# Asking for the selected operation in Admin Mode - Manage Doctors (Add, Delete, Edit or Display a Doctor) #
	print("Please choose one of the following options:")	
	print("0- Return Back.")
	print("1- Add a Doctor.")
	print("2- Delete a Doctor.")	
	print("3- Edit a Doctor.")	
	print("4- Display a Doctor.")	
	print("5- Display All Doctors.")	
	DoctorOption = str(input(">> "))
	print("")
	return DoctorOption
#################################################################################################
## A function to ask for the selected option in the admin mode - manage appointments ##
def AskForAdminModeBookAppointment():
	# Asking for the selected operation in Admin Mode - Manage Appointments (Book, Edit or Cancel an Appointment) #
	print("Please choose one of the following options:")	
	print("0- Return Back.")
	print("1- Book an Appointment.")
	print("2- Edit an Appointment.")	
	print("3- Delete an Appointment.")
	print("4- Display All Appointments.")
	AppointmentOption = str(input(">> "))
	print("")
	return AppointmentOption
#################################################################################################
## A function to ask for the selected mode in the user mode ##		
def AskForUserMode():
	# Asking for the selected mode in User Mode (View Departments, View Doctors, View Patients, View One Patient or View an Appointment) #
	print("Please choose one of the following features:")
	print("0- Return Back.")
	print("1- View All Hospital's Departments.")
	print("2- View All Hospital's Doctors in Details.")
	print("3- View All Hospital's Patients in Details.")
	print("4- View Only One Patient's Details Using Its ID.")
	print("5- View Booked Appointment(s) Using Doctor's ID.")
	UserMode = str(input(">> "))
	print("")	
	return UserMode
#################################################################################################

#################################################################################################
#### Admin Mode Handlers ####
#################################################################################################
## A function to add a patient ##			
def AdminModeAddPatientHandler():
	# Asking the user to choose the department #
	Department = AskForDepartment()	
	# Asking the user to enter the doctor's ID #
	DoctorID, Flag = AskForDoctorID(Department)
	if(Flag == True):
		# Asking the user to enter the patient's first name #
		PatientFirstName = str(input("Please enter the patient's first name: "))
		print("")
		# Asking the user to enter the patient's surname #	
		PatientSurname = str(input("Please enter the patient's surname: "))
		print("")
		# Asking the user to enter the patient's age #	
		PatientAge = AskForAge("Patient")
		# Asking the user to choose the patient's gender #
		PatientGender = AskForGender("Patient")
		# Asking the user to choose the patient's address #
		PatientAddress = AskForAddress("Patient")
		# Asking the user to enter the patient's phone number #
		PatientPhoneNumber = AskForPhoneNumber("Patient")
		# Asking the user to enter the patient's room number #
		PatientRoomNumber, Flag = AskForRoomNumber(Department)
		if(Flag == True):
			# Asking the user to choose the patient's condition #
			PatientCondition = AskForPatientCondition()
			# Asking the user to enter the patient's ID #
			PatientID = AddPersonID("Patient", Department)
			# Appending the input parameters into the patient's sheet #
			PatientsSheet.append([PatientID, 
								  PatientFirstName,
								  PatientSurname, 
								  PatientAge, 
								  GendersLookupTable(PatientGender), 
								  AddressesLookupTable(PatientAddress), 
								  PatientPhoneNumber, 
								  DepartmentsLookupTable(Department), 
								  DoctorID, 
								  PatientRoomNumber, 
								  PatientConditionsLookupTable(PatientCondition)])
			# Updating the excel file after modifications #
			WriteToExcel()
			print("Patient has successfully been added")
		else:
			pass
	else:
		pass
	print("")
#################################################################################################
## A function to delete a patient ##
def AdminModeDeletePatientHandler():
	while(True):
		# Asking the user to enter the patient's ID #
		print("Please enter the patient's ID:")
		print("(Format) -> '0-[Department ID (1 digit)]-[Patient ID (6 digits)]'")
		ID = str(input(">> "))
		# Checking if the entered ID is found in the database or not #
		Row = 1
		while(Row <= PatientsSheet.max_row): 
			if(ID == PatientsSheet[Row][0].value):
				break
			else:
				Row = Row + 1
		if(Row > PatientsSheet.max_row):
			print("ID is wrong or not found, try again")
		else:
			break
	# Deleting the selected patient from the database #
	PatientsSheet.delete_rows(Row, 1)
	# Updating the excel file after modifications #
	WriteToExcel()
	print("Patient has successfully been deleted")
	print("")
#################################################################################################
## A function to edit a patient ##
def AdminModeEditPatientHandler():
	while(True):
		# Asking the user to enter the patient's ID #
		print("Please enter the patient's ID:")
		print("(Format) -> '0-[Department ID (1 digit)]-[Patient ID (6 digits)]'")
		ID = str(input(">> "))
		# Checking if the entered ID is found in the database or not #
		Row = 1
		while(Row <= PatientsSheet.max_row): 
			if(ID == PatientsSheet[Row][0].value):
				break
			else:
				Row = Row + 1
		if(Row > PatientsSheet.max_row):
			print("ID is wrong or not found, try again")
		else:
			break
	print("")
	Department = DepartmentsLookupTable(PatientsSheet[Row][7].value, "StrToInt")
	# Asking the user to enter the doctor's ID #
	DoctorID, Flag = AskForDoctorID(Department)
	PatientsSheet[Row][8].value = DoctorID
	# Asking the user to enter the patient's first name #
	PatientsSheet[Row][1].value = str(input("Please enter the patient's first name: "))
	print("")
	# Asking the user to enter the patient's surname #	
	PatientsSheet[Row][2].value = str(input("Please enter the patient's surname: "))
	print("")
	# Asking the user to enter the patient's age #	
	PatientsSheet[Row][3].value = AskForAge("Patient")
	# Asking the user to choose the patient's gender #
	PatientsSheet[Row][4].value = GendersLookupTable(AskForGender("Patient"))
	# Asking the user to choose the patient's address #
	PatientsSheet[Row][5].value = AddressesLookupTable(AskForAddress("Patient"))
	# Asking the user to enter the patient's phone number #
	PatientsSheet[Row][6].value = AskForPhoneNumber("Patient")
	# Asking the user to enter the patient's room number #
	PatientRoomNumber, Flag = AskForRoomNumber(Department)
	PatientsSheet[Row][9].value = PatientRoomNumber
	# Asking the user to choose the patient's condition #
	PatientsSheet[Row][10].value = PatientConditionsLookupTable(AskForPatientCondition())
	# Updating the excel file after modifications #
	WriteToExcel()
	print("Patient has successfully been modified")
	print("")
#################################################################################################
## A function to display a patient ##
def AdminModeDisplayPatientHandler():
	while(True):
		# Asking the user to enter the patient's ID #
		print("Please enter the patient's ID:")
		print("(Format) -> '0-[Department ID (1 digit)]-[Patient ID (6 digits)]'")
		ID = str(input(">> "))
		# Checking if the entered ID is found in the database or not #
		Row = 1
		while(Row <= PatientsSheet.max_row): 
			if(ID == PatientsSheet[Row][0].value):
				PrintRow(PatientsSheet[1])
				PrintRow(PatientsSheet[Row])
				break
			else:
				Row = Row + 1
		if(Row > PatientsSheet.max_row):
			print("ID is wrong or not found, try again")
		else:
			break
	print("")
#################################################################################################
## A function to display all patients ##
def AdminModeDisplayPatientsHandler():
	PrintSheet(PatientsSheet)
	print("")
#################################################################################################
#################################################################################################
## A function to add a doctor ##			
def AdminModeAddDoctorHandler():
	# Asking the user to choose the department #
	Department = AskForDepartment()	
	# Asking the user to enter the doctor's first name #
	DoctorFirstName = str(input("Please enter the doctor's first name: "))
	print("")
	# Asking the user to enter the doctor's surname #	
	DoctorSurname = str(input("Please enter the doctor's surname: "))
	print("")
	# Asking the user to enter the doctor's age #	
	DoctorAge = AskForAge("Doctor")
	# Asking the user to choose the doctor's gender #
	DoctorGender = AskForGender("Doctor")
	# Asking the user to choose the doctor's address #
	DoctorAddress = AskForAddress("Doctor")
	# Asking the user to enter the doctor's phone number #
	DoctorPhoneNumber = AskForPhoneNumber("Doctor")
	# Asking the user to enter the doctor's ID #
	DoctorID = AddPersonID("Doctor", Department)
	# Appending the input parameters into the doctor's sheet #
	DoctorsSheet.append([DoctorID, 
						 DoctorFirstName,
						 DoctorSurname, 
						 DoctorAge, 
						 GendersLookupTable(DoctorGender), 
						 AddressesLookupTable(DoctorAddress), 
						 DoctorPhoneNumber, 
						 DepartmentsLookupTable(Department),
						 "08:00", 
						 "17:00"])
	# Updating the excel file after modifications #
	WriteToExcel()
	print("Doctor has successfully been added")
	print("")
#################################################################################################
## A function to delete a doctor ##
def AdminModeDeleteDoctorHandler():
	while(True):
		# Asking the user to enter the doctor's ID #
		print("Please enter the doctor's ID:")
		print("(Format) -> '1-[Department ID (1 digit)]-[Doctor ID (6 digits)]'")
		ID = str(input(">> "))
		# Checking if the entered ID is found in the database or not #
		Row = 1
		while(Row <= DoctorsSheet.max_row): 
			if(ID == DoctorsSheet[Row][0].value):
				break
			else:
				Row = Row + 1
		if(Row > DoctorsSheet.max_row):
			print("ID is wrong or not found, try again")
		else:
			break
	# Deleting the selected doctor from the database #
	DoctorsSheet.delete_rows(Row, 1)
	# Updating the excel file after modifications #
	WriteToExcel()
	print("Doctor has successfully been deleted")
	print("")
#################################################################################################
## A function to edit a doctor ##
def AdminModeEditDoctorHandler():
	while(True):
		# Asking the user to enter the doctor's ID #
		print("Please enter the doctor's ID:")
		print("(Format) -> '1-[Department ID (1 digit)]-[Doctor ID (6 digits)]'")
		ID = str(input(">> "))
		# Checking if the entered ID is found in the database or not #
		Row = 1
		while(Row <= DoctorsSheet.max_row): 
			if(ID == DoctorsSheet[Row][0].value):
				break
			else:
				Row = Row + 1
		if(Row > DoctorsSheet.max_row):
			print("ID is wrong or not found, try again")
		else:
			break
	print("")
	Department = DepartmentsLookupTable(DoctorsSheet[Row][7].value, "StrToInt")
	# Asking the user to enter the doctor's first name #
	DoctorsSheet[Row][1].value = str(input("Please enter the doctor's first name: "))
	print("")
	# Asking the user to enter the doctor's surname #	
	DoctorsSheet[Row][2].value = str(input("Please enter the doctor's surname: "))
	print("")
	# Asking the user to enter the doctor's age #	
	DoctorsSheet[Row][3].value = AskForAge("Doctor")
	# Asking the user to choose the doctor's gender #
	DoctorsSheet[Row][4].value = GendersLookupTable(AskForGender("Doctor"))
	# Asking the user to choose the doctor's address #
	DoctorsSheet[Row][5].value = AddressesLookupTable(AskForAddress("Doctor"))
	# Asking the user to enter the doctor's phone number #
	DoctorsSheet[Row][6].value = AskForPhoneNumber("Doctor")
	# Updating the excel file after modifications #
	WriteToExcel()
	print("Doctor has successfully been modified")
	print("")
#################################################################################################
## A function to display a doctor ##
def AdminModeDisplayDoctorHandler():
	while(True):
		# Asking the user to enter the doctor's ID #
		print("Please enter the doctor's ID:")
		print("(Format) -> '1-[Department ID (1 digit)]-[Doctor ID (6 digits)]'")
		ID = str(input(">> "))
		# Checking if the entered ID is found in the database or not #
		Row = 1
		while(Row <= DoctorsSheet.max_row): 
			if(ID == DoctorsSheet[Row][0].value):
				PrintRow(DoctorsSheet[1])
				PrintRow(DoctorsSheet[Row])
				break
			else:
				Row = Row + 1
		if(Row > DoctorsSheet.max_row):
			print("ID is wrong or not found, try again")
		else:
			break
	print("")
#################################################################################################
## A function to display all doctors ##
def AdminModeDisplayDoctorsHandler():
	PrintSheet(DoctorsSheet)
	print("")
#################################################################################################
#################################################################################################
## A function to book an appointment ##		
def AdminModeBookAppointmentHandler():
	# Asking the user to choose the department #
	Department = AskForDepartment()	
	# Asking the user to enter the doctor's ID #
	DoctorID, Flag = AskForDoctorID(Department)
	if(Flag == True):
		# Asking the user to enter the patient's first name #
		PatientFirstName = str(input("Please enter the patient's first name: "))
		print("")
		# Asking the user to enter the patient's surname #	
		PatientSurname = str(input("Please enter the patient's surname: "))
		print("")
		# Asking the user to enter the patient's age #	
		PatientAge = AskForAge("Patient")
		# Asking the user to choose the patient's gender #
		PatientGender = AskForGender("Patient")
		# Asking the user to choose the patient's address #
		PatientAddress = AskForAddress("Patient")
		# Asking the user to enter the patient's phone number #
		PatientPhoneNumber = AskForPhoneNumber("Patient")
		# Asking the user to enter the patient's ID #
		PatientID = AddPersonID("Patient", Department, "Appointment")
		# Asking the user to choose one of the available time slot(s) #
		AppointmentFrom, AppointmentTo = AskForAppointment(DoctorID)
		# Appending the input parameters into the patient's sheet #
		AppointmentsSheet.append([PatientID, 
								  PatientFirstName,
								  PatientSurname, 
								  PatientAge, 
								  GendersLookupTable(PatientGender), 
								  AddressesLookupTable(PatientAddress), 
								  PatientPhoneNumber, 
								  DepartmentsLookupTable(Department), 
								  DoctorID,
								  AppointmentFrom,
								  AppointmentTo])
		# Updating the excel file after modifications #
		WriteToExcel()
		print("Appointment has successfully been booked")
	else:
		pass
	print("")
#################################################################################################
## A function to edit an appointment ##			
def AdminModeEditAppointmentHandler():
	while(True):
		# Asking the user to enter the patient's ID #
		print("Please enter the patient's ID:")
		print("(Format) -> '2-[Department ID (1 digit)]-[Patient ID (6 digits)]'")
		ID = str(input(">> "))
		# Checking if the entered ID is found in the database or not #
		Row = 1
		while(Row <= AppointmentsSheet.max_row): 
			if(ID == AppointmentsSheet[Row][0].value):
				AppointmentsSheet[Row][9].value = '0'
				AppointmentsSheet[Row][10].value = '0'
				break
			else:
				Row = Row + 1
		if(Row > AppointmentsSheet.max_row):
			print("ID is wrong or not found, try again")
		else:
			break
	print("")
	Department = DepartmentsLookupTable(AppointmentsSheet[Row][7].value, "StrToInt")
	# Asking the user to enter the doctor's ID #
	DoctorID, Flag = AskForDoctorID(Department)
	AppointmentsSheet[Row][8].value = DoctorID
	# Asking the user to enter the patient's first name #
	AppointmentsSheet[Row][1].value = str(input("Please enter the patient's first name: "))
	print("")
	# Asking the user to enter the patient's surname #	
	AppointmentsSheet[Row][2].value = str(input("Please enter the patient's surname: "))
	print("")
	# Asking the user to enter the patient's age #	
	AppointmentsSheet[Row][3].value = AskForAge("Patient")
	# Asking the user to choose the patient's gender #
	AppointmentsSheet[Row][4].value = GendersLookupTable(AskForGender("Patient"))
	# Asking the user to choose the patient's address #
	AppointmentsSheet[Row][5].value = AddressesLookupTable(AskForAddress("Patient"))
	# Asking the user to enter the patient's phone number #
	AppointmentsSheet[Row][6].value = AskForPhoneNumber("Patient")
	# Asking the user to choose one of the available time slot(s) #
	AppointmentsSheet[Row][9].value, AppointmentsSheet[Row][10].value = AskForAppointment(AppointmentsSheet[Row][8].value)
	# Updating the excel file after modifications #
	WriteToExcel()
	print("Appointment has successfully been modified")
	print("")
#################################################################################################
## A function to cancel an appointment ##		
def AdminModeCancelAppointmentHandler():
	while(True):
		# Asking the user to enter the patient's ID #
		print("Please enter the patient's ID:")
		print("(Format) -> '2-[Department ID (1 digit)]-[Patient ID (6 digits)]'")
		ID = str(input(">> "))
		# Checking if the entered ID is found in the database or not #
		Row = 1
		while(Row <= AppointmentsSheet.max_row): 
			if(ID == AppointmentsSheet[Row][0].value):
				break
			else:
				Row = Row + 1
		if(Row > AppointmentsSheet.max_row):
			print("ID is wrong or not found, try again")
		else:
			break
	# Deleting the selected patient from the database #
	AppointmentsSheet.delete_rows(Row, 1)
	# Updating the excel file after modifications #
	WriteToExcel()
	print("Appointment has successfully been canceled")
	print("")
#################################################################################################
## A function to display all available appointments ##
def AdminModeDisplayAppointmentsHandler():
	PrintSheet(AppointmentsSheet)
	print("")
#################################################################################################
#################################################################################################

#################################################################################################
#### User Mode Handlers ####
#################################################################################################
## A function to view all the available departments in the hospital ##
def UserModeViewDepartmentsHandler():
	print("The available departments in the hospital are: ")
	PrintDepartments()
	print("")
#################################################################################################
## A function to view a list of all doctors in the hospital ##
def UserModeViewDoctorsHandler():
	PrintSheet(DoctorsSheet)
	print("")
#################################################################################################
## A function to view a list of all patients in the hospital ##
def UserModeViewPatientsHandler():
	PrintSheet(PatientsSheet)
	print("")
#################################################################################################
## A function to view the information of one patient in the hospital ##
def UserModeViewOnePatientHandler():
	while(True):
		# Asking the user to enter the patient's ID #
		print("Please enter the patient's ID:")
		print("(Format) -> '0-[Department ID (1 digit)]-[Patient ID (6 digits)]'")
		ID = str(input(">> "))
		# Checking if the entered ID is found in the database or not #
		Row = 1
		while(Row <= PatientsSheet.max_row): 
			if(ID == PatientsSheet[Row][0].value):
				PrintRow(PatientsSheet[1])
				PrintRow(PatientsSheet[Row])
				break
			else:
				Row = Row + 1
		if(Row > PatientsSheet.max_row):
			print("ID is wrong or not found, try again")
		else:
			break
	print("")
#################################################################################################
## A function to view the information of a booked appointment in the hospital ##
def UserModeViewAppointmentHandler():
	Flag = False
	while(True):
		# Asking the user to enter the doctor's ID #
		print("Please enter the doctor's ID:")
		print("(Format) -> '1-[Department ID (1 digit)]-[Doctor ID (6 digits)]'")
		ID = str(input(">> "))
		# Checking if the entered ID is found in the database or not #
		Row = 1
		while(Row <= AppointmentsSheet.max_row): 
			if(ID == AppointmentsSheet[Row][8].value):
				if(Flag == False):
					PrintRow(AppointmentsSheet[1])
				else:
					pass
				PrintRow(AppointmentsSheet[Row])
				Flag = True
			else:
				pass
			Row = Row + 1
		if(Flag == False):
			print("ID is wrong, not found or no appointments available for this doctor, try again")
		else:
			break
	print("")
#################################################################################################
#################################################################################################

#################################################################################################
### Error Handlers ###
#################################################################################################
## A function to display that a wrong entry has been detected ##		
def ErrorHandler():
	print("Wrong entry, try again")
	print("")
#################################################################################################
## A function to display that the system is locked ##
def LockSystem():
	print("				|*********************************************************************|")
	print("				|****************** ACCESS DENIED, SYSTEM IS LOCKED ******************|")
	print("				|*********************************************************************|")
	print("")
#################################################################################################


#################################################################################################
#################################################################################################
##### PRIVATE FUNCTIONS #####
#################################################################################################
#################################################################################################

#################################################################################################
### Excel File Handlers ###
#################################################################################################
## A function to read from an excel file which contains all hospital information ##
def ReadFromExcel():
	global HospitalSystemFile
	global PatientsSheet
	global DoctorsSheet
	global AppointmentsSheet
	HospitalSystemFile = ExcelHandler.load_workbook("Hospital System.xlsx")
	PatientsSheet = HospitalSystemFile.get_sheet_by_name("Patients")
	DoctorsSheet = HospitalSystemFile.get_sheet_by_name("Doctors")
	AppointmentsSheet = HospitalSystemFile.get_sheet_by_name("Appointments")
#################################################################################################
## A function to write to an excel file which contains all hospital information ##
def WriteToExcel():
	while(True):
		try:
			HospitalSystemFile.save("Hospital System.xlsx")
		except:
			print("Excel file is already opened and inaccessible.")
			input(str("Please close the file and enter any value to retry: "))
			print("")
		else:
			break
#################################################################################################

#################################################################################################
### User Input Handlers ###
#################################################################################################
## A function to ask the user to choose the department ##
def AskForDepartment():
	while(True):
		# Asking the user to choose the department #
		print("Please choose one of the following departments:")
		PrintDepartments()
		Department = str(input(">> "))
		# Checking if the entered department is valid or not #
		try:
			if((int(Department) < 10) and (int(Department) > -1)):
				pass
		except:
			# Wrong entry #
			print("Wrong entry, try again")
		else:
			if((int(Department) < 10) and (int(Department) > -1)):
				break
			else:
				# Wrong entry #
				print("Wrong entry, try again")
	print("")
	return Department
#################################################################################################
## A function to ask the user to enter the patient/doctor's age ##	
def AskForAge(Person):
	while(True):
		# Asking the user to enter the patient/doctor's age #	
		Age = str(input("Please enter the " + Person.lower() + "'s age: "))
		# Checking if the entered age is valid or not #
		try:
			if((int(Age) < 100) and (int(Age) > 0)):
				pass
		except:
			# Wrong entry #
			print("Wrong entry, try again")
		else:
			if((int(Age) < 100) and (int(Age) > 0)):
				break
			else:
				# Wrong entry #
				print("Wrong entry, try again")
	print("")
	return Age
#################################################################################################
## A function to ask the user to choose the patient/doctor's gender ##
def AskForGender(Person):
	while(True):
		# Asking the user to choose the patient/doctor's gender #
		print("Please choose the " + Person.lower() + "'s gender:")
		PrintGenders()	
		Gender = str(input(">> "))
		try:
			if((int(Gender) < 3) and (int(Gender) > 0)):
				pass
		except:
			# Wrong entry #
			print("Wrong entry, try again")
		else:
			if((int(Gender) < 3) and (int(Gender) > 0)):
				break
			else:
				# Wrong entry #
				print("Wrong entry, try again")
	print("")
	return Gender
#################################################################################################
## A function to ask the user to choose the patient/doctor's address ##
def AskForAddress(Person):
	while(True):
		# Asking the user to choose the patient/doctors's address #
		print("Please choose the " + Person.lower() + "'s address from the following addresses:")
		PrintAddresses()
		Address = str(input(">> "))
		# Checking if the entered address is valid or not #
		try:
			if((int(Address) < 25) and (int(Address) > 0)):
				pass
		except:
			# Wrong entry #
			print("Wrong entry, try again")
		else:
			if((int(Address) < 25) and (int(Address) > 0)):
				break
			else:
				# Wrong entry #
				print("Wrong entry, try again")
	print("")
	return Address
#################################################################################################
## A function to ask the user to enter the patient/doctor's phone number ##
def AskForPhoneNumber(Person):
	while(True):
		# Asking the user to enter the patient/doctor's phone number #
		PhoneNumber = str(input("Please enter the " + Person.lower() + "'s phone number: "))
		# Checking if the length of the input equals 11 (as phone number) #
		while(len(PhoneNumber) != 11):
			# Wrong entry #
			print("Wrong entry, try again")
			# Asking the user to enter the patient/doctor's phone number #
			PhoneNumber = str(input("Please enter the " + Person.lower() + "'s phone number: "))
		# Checking if the entered number is a valid phone number #	
		try:
			if((int(PhoneNumber[3:]) < 100000000) and (int(PhoneNumber[3:]) > -1)):
				pass
		except:
			# Wrong entry #
			print("Wrong entry, try again")
		else:
			if((PhoneNumber[0] == '0') and (PhoneNumber[1] == '1') and ((PhoneNumber[2] == '0') or (PhoneNumber[2] == '1') or (PhoneNumber[2] == '2') or (PhoneNumber[2] == '5')) and (int(PhoneNumber[3:]) < 100000000) and (int(PhoneNumber[3:]) > -1)):
				break
			else:
				# Wrong entry #
				print("Wrong entry, try again")
	print("")
	return PhoneNumber
#################################################################################################
## A function to ask the user to enter the patient/doctor's ID ##
def AddPersonID(Person, Department, Type = "Residence"):
	while(True):
		while(True):
			# Asking the user to enter the patient/doctor's ID #
			PersonID = str(input("Please enter the " + Person.lower() + "'s ID (from 000000 to 999999): "))
			# Checking if the entered ID is valid or not #
			while(len(PersonID) != 6):
				# Wrong entry #
				print("Wrong entry, try again")
				# Asking the user to enter the patient/doctor's ID #
				PersonID = str(input("Please enter the " + Person.lower() + "'s ID (from 000000 to 999999): "))
			try:
				if((int(PersonID) < 1000000) and (int(PersonID) > -1)):
					pass
			except:
				# Wrong entry #
				print("Wrong entry, try again")
			else:
				break
		if((Person == "Patient") and (Type == "Residence")):
			# Generating the full Patient ID #
			PersonID = "0-" + Department + '-' + PersonID
			# Checking if the entered ID is found in the database or not #
			Row = 1
			while(Row <= PatientsSheet.max_row): 
				if(PersonID == PatientsSheet[Row][0].value):
					break
				else:
					Row = Row + 1
			# Printing that the entered ID is wrong or repeated and ask the user to enter the patient's ID one more time #
			if(Row <= PatientsSheet.max_row):
				print("Entry is rejected, ID is already found")
			else:
				break
		elif((Person == "Patient") and (Type == "Appointment")):
			# Generating the full Patient ID #
			PersonID = "2-" + Department + '-' + PersonID
			# Checking if the entered ID is found in the database or not #
			Row = 1
			while(Row <= AppointmentsSheet.max_row): 
				if(PersonID == AppointmentsSheet[Row][0].value):
					break
				else:
					Row = Row + 1
			# Printing that the entered ID is wrong or repeated and ask the user to enter the patient's ID one more time #
			if(Row <= AppointmentsSheet.max_row):
				print("Entry is rejected, ID is already found")
			else:
				break
		else:
			# Generating the full Doctor ID #
			PersonID = "1-" + Department + '-' + PersonID
			# Checking if the entered ID is found in the database or not #
			Row = 1
			while(Row <= DoctorsSheet.max_row): 
				if(PersonID == DoctorsSheet[Row][0].value):
					break
				else:
					Row = Row + 1
			# Printing that the entered ID is wrong or repeated and ask the user to enter the doctor's ID one more time #
			if(Row <= DoctorsSheet.max_row):
				print("Entry is rejected, ID is already found")
			else:
				break
	print("")
	return PersonID
#################################################################################################
## A function to ask the user to enter the patient's room number ##
def AskForRoomNumber(Department):
	PatientRoomNumber = '0'
	while(True):
		while(True):
			# Asking the user to enter the patient's room number #
			print("Please enter the patient's room number (from 0 to 9):")
			RoomsFlag = PrintRooms(Department)
			if(RoomsFlag == True):
				PatientRoomNumber = str(input(">> "))			
				# Checking if the entered room is already taken or not #
				Row = 1
				Flag = False
				while(Row <= PatientsSheet.max_row): 
					if((Department == PatientsSheet[Row][9].value[0]) and (PatientRoomNumber == PatientsSheet[Row][9].value[1])):
						Flag = True
					else:
						pass
					Row = Row + 1
				if(Flag == True):
					print("Room is already taken, try again")
				else:
					break
			else:
				break
		if(RoomsFlag == True):
			try:
				if((int(PatientRoomNumber) < 10) and (int(PatientRoomNumber) > -1)):
					pass
			except:
				# Wrong entry #
				print("Wrong entry, try again")
			else:
				if((int(PatientRoomNumber) < 10) and (int(PatientRoomNumber) > -1)):
					PatientRoomNumber = Department + PatientRoomNumber
					break
				else:
					# Wrong entry #
					print("Wrong entry, try again")
		else:
			break
	print("")
	return PatientRoomNumber, RoomsFlag
#################################################################################################
## A function to ask the user to choose the patient's condition ##
def AskForPatientCondition():
	while(True):
		# Asking the user to choose the patient's condition #
		print("Please choose the patient's condition from the following options:")
		PrintPatientConditions()
		PatientCondition = str(input(">> "))
		# Checking if the entered patient condition is valid or not #
		try:
			if((int(PatientCondition) < 5) and (int(PatientCondition) > 0)):
				pass
		except:
			# Wrong entry #
			print("Wrong entry, try again")
		else:
			if((int(PatientCondition) < 5) and (int(PatientCondition) > 0)):
				break
			else:
				# Wrong entry #
				print("Wrong entry, try again")
	print("")
	return PatientCondition
#################################################################################################
## A function to ask the user to enter the doctor's ID ##
def AskForDoctorID(Department):
	DoctorID = '0'
	while(True):
		# Asking the user to enter the doctor's ID #
		Flag = PrintDoctors(Department)
		if(Flag == True):
			DoctorID = str(input(">> "))
			# Checking if the entered ID is found in the database or not #
			Row = 1
			while(Row <= DoctorsSheet.max_row): 
				if(DoctorID == DoctorsSheet[Row][0].value):
					break
				else:
					Row = Row + 1
			if(Row > DoctorsSheet.max_row):
				print("ID is wrong or not found, try again")
			else:
				break
		else:
			break
	print("")
	return DoctorID, Flag	
#################################################################################################
## A function to ask the user to choose the suitable appointment time slots ##
def AskForAppointment(DoctorID):
	# Fetching the row value that corresponds the input Doctor ID #
	Row = 1
	while(Row <= DoctorsSheet.max_row): 
		if(DoctorID == DoctorsSheet[Row][0].value):
			break
		else:
			Row = Row + 1
	# Saving the values of start and end times of the selected doctor #
	StartTime = int(DoctorsSheet[Row][8].value[0:2])
	EndTime = int(DoctorsSheet[Row][9].value[0:2])
	# Fetching all unavailable time slots that corresponds the selected #
	Row = 1
	BusyTimeFrom = list()
	while(Row <= AppointmentsSheet.max_row): 
		if(DoctorID == AppointmentsSheet[Row][8].value):
			BusyTimeFrom.append(int(AppointmentsSheet[Row][9].value[0:2]))
		else:
			pass
		Row = Row + 1
	# Filtering the available time slots for the selected doctor
	PrintIndex = 1
	TimeSlots = list()
	print("Please choose one of the available time slot(s):")
	while(StartTime < EndTime):
		ListIndex = 0
		# Checking if there are any appointments already booked in this time slot #
		while(ListIndex < len(BusyTimeFrom)): 
			if(StartTime == BusyTimeFrom[ListIndex]):		
				break
			else:
				ListIndex = ListIndex + 1
		# Printing the current time slot if it is available #
		if(ListIndex == len(BusyTimeFrom)):
			print(str(PrintIndex) + "- From " + str(StartTime) + ":00 to " + str(StartTime + 1) + ":00.")
			TimeSlots.append(StartTime)
			PrintIndex = PrintIndex + 1
		else:
			pass
		StartTime = StartTime + 1
	# Asking the user to choose from the available time slot(s) #
	while(True):
		# Asking the user to enter the suitable time slot #
		Index = str(input(">> "))
		try:
			# Checking if the entered value is valid or not #
			if((int(Index) < PrintIndex) and (int(Index) > 0)):
				pass
		except:
			# Wrong entry #
			print("Wrong entry, try again")
		else:
			if((int(Index) < PrintIndex) and (int(Index) > 0)):
				# Saving the values of the starting and the ending time of the appointment # 
				AppointmentFrom = str(TimeSlots[int(Index) - 1]) + ":00"
				AppointmentTo = str(TimeSlots[int(Index) - 1] + 1) + ":00"
				break
			else:
				# Wrong entry #
				print("Wrong entry, try again")
	return AppointmentFrom, AppointmentTo
#################################################################################################

#################################################################################################
### Printing Functions ###
#################################################################################################
## A function to print the content of a specific sheet ##		
def PrintSheet(Sheet):
	for Row in Sheet.iter_rows(max_row = Sheet.max_row):
		for Cell in Row:
			print('{:^20s}'.format(str(Cell.value)), end = '')
		print()
#################################################################################################
## A function to print the content of an entire row in a specific sheet ##		
def PrintRow(SheetRow):
	for Cell in SheetRow:
		print('{:^20s}'.format(str(Cell.value)), end = '')
	print()
#################################################################################################
## A function to print all hospital's departments ##
def PrintDepartments():
	print(" 0- Cardilogy.")
	print(" 1- Ear, Nose and Throat.")
	print(" 2- Gastroenterology.")
	print(" 3- General Surgery.")
	print(" 4- Gynecology.")
	print(" 5- Neurology.")
	print(" 6- Obstetrics.")
	print(" 7- Orthopedics.")
	print(" 8- Opthalmology.")
	print(" 9- Urology.")
#################################################################################################
## A function to print genders ## 
def PrintGenders():
	print("1- Male.")
	print("2- Female.")
#################################################################################################
## A function to print all available addresses ##
def PrintAddresses():
	print("1-	Abo Qeer.")
	print("2-	Al Assafra.")
	print("3-	Al Azarita.")
	print("4-	Al Ibrahimiyya.")
	print("5-	Al Maamora.")
	print("6-	Al Mandara.")
	print("7-	Al Manshia.")
	print("8-	Al Montazah.")
	print("9-	Al Seyouf.")
	print("10-	Al Shatby.")
	print("11-	Bahary.")
	print("12-	Bulkley.")
	print("13-	Camp Caesar.")
	print("14-	Cleopatra.")
	print("15-	Fleming.")
	print("16-	Gianaklis.")
	print("17-	Gleem.")
	print("18-	Mahattet Al Ramleh.")
	print("19-	Mustafa Kamil.")
	print("20-	Roushdy.")
	print("21-	San Stefano.")
	print("22-	Sidi Beshr.")
	print("23-	Sidi Gaber.")
	print("24-	Sporting.")
#################################################################################################
## A function to print all available patient conditions ##
def PrintPatientConditions():
	print("1- Good.")
	print("2- Fair.")
	print("3- Serious.")
	print("4- Critcal.")
#################################################################################################
## A function to print all available doctors within the selected department ##
def PrintDoctors(Department):
	Row = 1
	Flag = False
	while(Row <= DoctorsSheet.max_row):
		if(DepartmentsLookupTable(Department) == DoctorsSheet[Row][7].value):
			if(Flag == False):
				print("Please choose one of the following available doctor(s) by writing its appropiate ID:")
				PrintRow(DoctorsSheet[1])
				Flag = True
			else:
				pass
			PrintRow(DoctorsSheet[Row])
		else:
			pass
		Row = Row + 1
	if(Flag == False):
		print("No doctors are available for this department. Process is terminated.")
		print("")
	else:
		pass
	return Flag
#################################################################################################
## A function to print all available rooms within the selected department ##
def PrintRooms(Department):
	Row = 1
	Flag = True
	Rooms = list()
	while(Row <= PatientsSheet.max_row):
		if(DepartmentsLookupTable(Department) == PatientsSheet[Row][7].value):
			Rooms.append(PatientsSheet[Row][9].value[1])
		else:
			pass
		Row = Row + 1
	if(len(Rooms) == 0):
		print("All rooms are available for selection.")
	elif(len(Rooms) < 10):
		print("This/These room(s) are already taken: " + str(Rooms))
	else:
		print("All rooms are already taken. Process is terminated")
		Flag = False
	return Flag
#################################################################################################

#################################################################################################
### Lookup Table Functions ###
#################################################################################################
## A function to convert departments from intergers to strings or vice versa ##
def DepartmentsLookupTable(Department, Type = "IntToStr"):
	# Returning the string for the equiavlent number or vice versa #
	if(Type == "StrToInt"):
		if(Department == "Cardilogy"):
			Department = '0'
		elif(Department == "Ear, Nose and Throat"):
			Department = '1'
		elif(Department == "Gastroenterology"):
			Department = '2'
		elif(Department == "General Surgery"):
			Department = '3'
		elif(Department == "Gynecology"):
			Department = '4'
		elif(Department == "Neurology"):
			Department = '5'
		elif(Department == "Obstetrics"):
			Department = '6'
		elif(Department == "Orthopedics"):
			Department = '7'
		elif(Department == "Opthalmology"):
			Department = '8'
		else:
			Department = '9'	
	else:
		if(Department == '0'):
			Department = "Cardilogy"
		elif(Department == '1'):
			Department = "Ear, Nose and Throat"
		elif(Department == '2'):
			Department = "Gastroenterology"
		elif(Department == '3'):
			Department = "General Surgery"
		elif(Department == '4'):
			Department = "Gynecology"
		elif(Department == '5'):
			Department = "Neurology"
		elif(Department == '6'):
			Department = "Obstetrics"
		elif(Department == '7'):
			Department = "Orthopedics"
		elif(Department == '8'):
			Department = "Opthalmology"
		else:
			Department = "Urology"
	return Department
#################################################################################################
## A function to convert genders from integers to strings ##
def GendersLookupTable(Gender):
	# Returning the string for the equiavlent number #
	if(Gender == '1'):
		Gender = "Male"
	else:
		Gender = "Female"
	return Gender
#################################################################################################
## A function to convert addresses from integers to strings ##
def AddressesLookupTable(Address):
	# Returning the string for the equiavlent number #
	if(Address == '1'):
		Address = "Abo Qeer"
	elif(Address == '2'):
		Address = "Al Assafra"
	elif(Address == '3'):
		Address = "Al Azarita"
	elif(Address == '4'):
		Address = "Al Ibrahimiyya"
	elif(Address == '5'):
		Address = "Al Maamora"
	elif(Address == '6'):
		Address = "Al Mandara"
	elif(Address == '7'):
		Address = "Al Manshia"
	elif(Address == '8'):
		Address = "Al Montazah"
	elif(Address == '9'):
		Address = "Al Seyouf"
	elif(Address == '10'):
		Address = "Al Shatby"
	elif(Address == '11'):
		Address = "Bahary"
	elif(Address == '12'):
		Address = "Bulkley"
	elif(Address == '13'):
		Address = "Camp Caesar"
	elif(Address == '14'):
		Address = "Cleopatra"
	elif(Address == '15'):
		Address = "Fleming"
	elif(Address == '16'):
		Address = "Gianaklis"
	elif(Address == '17'):
		Address = "Gleem"
	elif(Address == '18'):
		Address = "Mahattet Al Ramleh"
	elif(Address == '19'):
		Address = "Mustafa Kamil"
	elif(Address == '20'):
		Address = "Roushdy"
	elif(Address == '21'):
		Address = "San Stefano"
	elif(Address == '22'):
		Address = "Sidi Beshr"
	elif(Address == '23'):
		Address = "Sidi Gaber"
	else:
		Address = "Sporting"
	return Address
#################################################################################################
## A function to convert patient conditions from integers to strings ##
def PatientConditionsLookupTable(PatientCondition):
	# Returning the string for the equiavlent number #
	if(PatientCondition == '1'):
		PatientCondition = "Good"
	elif(PatientCondition == '2'):
		PatientCondition = "Fair"
	elif(PatientCondition == '3'):
		PatientCondition = "Serious"
	else:
		PatientCondition = "Critcal"
	return PatientCondition
#################################################################################################
